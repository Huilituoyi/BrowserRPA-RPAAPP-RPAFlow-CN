# -*- coding: utf-8 -*-
"""Excel 导出单元测试。"""
import os

import pytest
from openpyxl import load_workbook

from core.data.excel_exporter import export


class TestExport:
    """export() 函数。"""

    def test_normal_data(self, tmp_data_dir):
        rows = [
            {"姓名": "张三", "年龄": 25, "城市": "北京"},
            {"姓名": "李四", "年龄": 30, "城市": "上海"},
        ]
        path = os.path.join(tmp_data_dir, "test.xlsx")
        result = export(rows, path)

        assert os.path.exists(result)
        wb = load_workbook(path)
        ws = wb.active
        # 表头
        assert ws.cell(1, 1).value == "姓名"
        assert ws.cell(1, 2).value == "年龄"
        # 数据行
        assert ws.cell(2, 1).value == "张三"
        assert ws.cell(3, 2).value == 30  # 数字保持数字类型
        wb.close()

    def test_empty_rows(self, tmp_data_dir):
        path = os.path.join(tmp_data_dir, "empty.xlsx")
        result = export([], path)
        assert os.path.exists(result)

    def test_empty_with_headers(self, tmp_data_dir):
        path = os.path.join(tmp_data_dir, "headers_only.xlsx")
        result = export([], path, headers=["A", "B", "C"])
        assert os.path.exists(result)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 3).value == "C"
        wb.close()

    def test_custom_headers_order(self, tmp_data_dir):
        rows = [{"a": 1, "b": 2, "c": 3}]
        path = os.path.join(tmp_data_dir, "custom.xlsx")
        export(rows, path, headers=["c", "a", "b"])
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "c"
        assert ws.cell(1, 3).value == "b"
        wb.close()

    def test_none_value(self, tmp_data_dir):
        """None 值导出不报错；openpyxl 将空单元格读回为 None，等价于空串。"""
        rows = [{"a": "x", "b": None}]
        path = os.path.join(tmp_data_dir, "none.xlsx")
        export(rows, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 2).value in (None, "")
        wb.close()

    def test_column_width_set(self, tmp_data_dir):
        """列宽被设置。"""
        rows = [{"name": "a"}]
        path = os.path.join(tmp_data_dir, "width.xlsx")
        export(rows, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["A"].width >= 8
        wb.close()
