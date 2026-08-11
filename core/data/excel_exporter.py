# -*- coding: utf-8 -*-
"""
Excel 导出：把 list[dict] 数据导出为 .xlsx（openpyxl）。
自动以第一条记录的 key 作为表头，并做简单的列宽自适应。
"""
import os
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.logging.logger import get_logger

log = get_logger("excel")


def export(rows: List[Dict], path: str, sheet_name: str = "Sheet1",
           headers: Optional[List[str]] = None) -> str:
    """
    :param rows: 数据行（list[dict]）
    :param path: 输出 .xlsx 路径
    :param sheet_name: 工作表名
    :param headers: 自定义表头顺序；为空时取 rows[0] 的 key
    :return: 实际保存路径
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel 表名最长 31

    if not rows and not headers:
        wb.save(path)
        log.info("数据为空，已生成空白 Excel：%s", path)
        return path

    if not headers:
        headers = list(rows[0].keys()) if rows else []

    ws.append(headers)
    for r in rows:
        ws.append([_to_cell(r.get(h, "")) for h in headers])

    # 简单列宽自适应（按表头与样本长度估算）
    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for r in rows[:200]:  # 仅采样前 200 行
            v = r.get(h, "")
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 8), 60)

    wb.save(path)
    log.info("已导出 Excel：%s（%d 行 × %d 列）", path, len(rows), len(headers))
    return path


def _to_cell(v):
    """Excel 单元格值规整：保留基础类型。"""
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return ""
    return str(v)
